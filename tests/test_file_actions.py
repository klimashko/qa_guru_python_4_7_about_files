from PyPDF2 import PdfReader
import os
import csv
from zipfile import ZipFile
import codecs
from openpyxl import load_workbook




def test_zip_csv(clean_resources):
    with open('addresses_1.csv') as file:
        rows = [row for row in csv.reader(file)]
        first_row = rows[0]
        num_rows = len(rows)


    with ZipFile(os.path.join(os.path.dirname(os.path.abspath(__file__)), "resources", "addresses_1_csv.zip"), "w") as my_zip:
        my_zip.write('addresses_1.csv')
    with ZipFile('resources/addresses_1_csv.zip') as myzip:
        with myzip.open('addresses_1.csv', 'r') as csv_file:
            rows_from_zip = [row for row in csv.reader(codecs.iterdecode(csv_file, 'utf-8'))]
            first_row_from_zip = rows_from_zip[0]
            num_rows_from_zip = len(rows_from_zip)


    assert first_row_from_zip == first_row, 'First row not matched'
    assert num_rows_from_zip == num_rows, 'Rows numbers in csv files not matched'




def test_zip_xlsx(clean_resources):
    with open('sample2.xlsx') as file:
        cell = load_workbook('sample2.xlsx').active.cell(row=1, column=2).value
        worksheet = load_workbook('sample2.xlsx').active
        num_rows = 0
        for row in worksheet.iter_rows():
            if any(cell.value for cell in row):
                num_rows += 1


    with ZipFile(os.path.join(os.path.dirname(os.path.abspath(__file__)), "resources", "sample2_xlsx.zip"), "w") as new_zip:
         new_zip.write("sample2.xlsx")
    with ZipFile("resources/sample2_xlsx.zip") as myzip:
        with myzip.open('sample2.xlsx', 'r') as xlsx_file:
            cell_from_zip = load_workbook(xlsx_file).active.cell(row=1, column=2).value
            worksheet = load_workbook(xlsx_file).active
            num_rows_from_zip = 0
            for row in worksheet.iter_rows():
                if any(cell.value for cell in row):
                    num_rows_from_zip += 1

    assert cell == cell_from_zip, 'Cell content not matched'
    assert num_rows == num_rows_from_zip, 'Row numbers not matched'




def test_zip_pdf(clean_resources):
    num_pages = len(PdfReader('example3.pdf').pages)
    text_1page = PdfReader('example3.pdf').pages[0].extract_text()


    with ZipFile(os.path.join(os.path.dirname(os.path.abspath(__file__)), "resources", "example3_pdf.zip"), "w") as my_zip:
        my_zip.write('example3.pdf')
    with ZipFile('resources/example3_pdf.zip') as my_zip:
        with my_zip.open('example3.pdf', 'r') as pdf_file:
            num_pages_from_zip = len(PdfReader(pdf_file).pages)
            text_1page_from_zip = PdfReader(pdf_file).pages[0].extract_text()


    assert num_pages_from_zip == num_pages, "Number of pages does not match"
    assert text_1page_from_zip == text_1page, "Beginning text not the same"



